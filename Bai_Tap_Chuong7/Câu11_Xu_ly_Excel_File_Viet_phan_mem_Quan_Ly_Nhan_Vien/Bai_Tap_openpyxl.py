import openpyxl
from openpyxl import Workbook

file_path = "nhanvien.xlsx"

# ====== HÀM KHỞI TẠO FILE EXCEL NẾU CHƯA CÓ ======
def tao_file_moi():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "NhanVien"
        ws.append(["STT", "Mã", "Tên", "Tuổi"])
        wb.save(file_path)
        print("Đã tạo file Excel mới.")
    except Exception as e:
        print("Lỗi khi tạo file:", e)

# ====== HÀM GHI NHÂN VIÊN VÀO FILE ======
def them_nhan_vien(ma, ten, tuoi):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        stt = ws.max_row  # dòng kế tiếp
        ws.append([stt, ma, ten, tuoi])
        wb.save(file_path)
        print("Đã thêm nhân viên thành công!")
    except FileNotFoundError:
        print("File chưa tồn tại. Vui lòng tạo mới bằng cách chạy 'tao_file_moi()'.")

# ====== HÀM ĐỌC DANH SÁCH NHÂN VIÊN ======
def doc_danh_sach():
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        print("\n--- DANH SÁCH NHÂN VIÊN ---")
        for row in ws.iter_rows(min_row=2, values_only=True):
            print(f"STT: {row[0]} | Mã: {row[1]} | Tên: {row[2]} | Tuổi: {row[3]}")
    except FileNotFoundError:
        print("File không tồn tại!")

# ====== HÀM SẮP XẾP NHÂN VIÊN THEO TUỔI ======
def sap_xep_theo_tuoi():
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(list(row))

        data.sort(key=lambda x: x[3])  # sắp xếp theo cột tuổi

        ws.delete_rows(2, ws.max_row)  # xóa dữ liệu cũ
        for i, nv in enumerate(data, start=1):
            ws.append([i, nv[1], nv[2], nv[3]])

        wb.save(file_path)
        print("Đã sắp xếp nhân viên theo tuổi tăng dần!")
    except FileNotFoundError:
        print("File không tồn tại!")

# ====== MENU CHƯƠNG TRÌNH ======
def menu():
    while True:
        print("\n===== QUẢN LÝ NHÂN VIÊN =====")
        print("1. Tạo file Excel mới")
        print("2. Thêm nhân viên")
        print("3. Xem danh sách nhân viên")
        print("4. Sắp xếp theo tuổi tăng dần")
        print("0. Thoát")
        chon = input("Nhập lựa chọn: ")

        if chon == "1":
            tao_file_moi()
        elif chon == "2":
            ma = input("Nhập mã NV: ")
            ten = input("Nhập tên NV: ")
            tuoi = int(input("Nhập tuổi NV: "))
            them_nhan_vien(ma, ten, tuoi)
        elif chon == "3":
            doc_danh_sach()
        elif chon == "4":
            sap_xep_theo_tuoi()
        elif chon == "0":
            print("Thoát chương trình.")
            break
        else:
            print("Lựa chọn không hợp lệ!")

# ====== CHẠY CHƯƠNG TRÌNH ======
menu()